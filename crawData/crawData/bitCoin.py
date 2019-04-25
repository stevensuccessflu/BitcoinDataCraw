import requests
import json
import time,os
import xlrd,xlwt
from openpyxl import load_workbook
from bs4 import BeautifulSoup

#获取比特币价格
def get_allBitCoinPrice():
    response = requests.get('https://api.gbi.news/coin-points?type=all&coin=BTC')
    jsonstr = json.loads(response.text)
    price_list = jsonstr['data']['points']
    path = 'D:\data\\bitcoin\\bitcoin.csv'
    #step1 初始化工作表
    f = xlwt.Workbook()
    #step2 创建表单
    sheet1 = f.add_sheet(r'btcprice', cell_overwrite_ok=True)
    #step3 写入数据
    for i in range(len(price_list)):
        timestap = int(price_list[i][0]/1000)
        timestamp = time.strftime("%Y-%m-%d", time.localtime(timestap))
        price = round(price_list[i][1],2)
        sheet1.write(i,0,timestamp)
        sheet1.write(i, 1, price)
        f.save(path)

def get_allBitCoinGBI():
    response = requests.get('https://api.gbi.news/gbi?type=all')
    jsonstr = json.loads(response.text)
    gbi_list = jsonstr['data']['points']
    path = 'D:\data\\bitcoin\\bitcoinGBI.csv'
    # step1 初始化工作表
    f = xlwt.Workbook()
    # step2 创建表单
    sheet1 = f.add_sheet(r'GBI', cell_overwrite_ok=True)
    for i in range(len(gbi_list)):
        timestap = int(gbi_list[i][0] / 1000)
        timestamp = time.strftime("%Y-%m-%d", time.localtime(timestap))
        price = round(gbi_list[i][1], 2)
        sheet1.write(i, 0, timestamp)
        sheet1.write(i, 1, price)
        f.save(path)

#场外usdt溢价
def get_foxOTC():
    response = requests.get('https://api.gbi.news/foi?type=all')
    jsonstr = json.loads(response.text)
    fox_list = jsonstr['data']['points']
    path = 'D:\data\\bitcoin\\bitcoinFox1.csv'
    # step1 初始化工作表
    f = xlwt.Workbook()
    # step2 创建表单
    sheet1 = f.add_sheet(r'fox', cell_overwrite_ok=True)
    for i in range(len(fox_list)):
        timestap = int(fox_list[i]['timestamp']/1000)
        timestamp = time.strftime("%Y-%m-%d", time.localtime(timestap))
        buyRate = fox_list[i]['buyRate']
        if buyRate[0] =='+':
            buyRate = round(float(buyRate[1:-1])*100,2)
        else:
            buyRate = round(-float(buyRate[1:-1])*100,2)
        sheet1.write(i, 0, timestamp)
        sheet1.write(i, 1, buyRate)
        f.save(path)

#获取BTC,LTC,BCH,EOS,ETH的换手率
def get_tournover_rate():
    coin_list = ['BTC','ETH','EOS','XRP','LTC','BCH']
    path = 'D:\data\\bitcoin\\turn.xlsx'
    base_url = 'https://bitkan.com/zh/currencies/'

    if os.path.exists(path):
        #1、打开文件
        workbook = load_workbook(path)
        for i in range(len(coin_list)):
            response = requests.get(base_url+coin_list[i])
            soup = BeautifulSoup(response.text,"html.parser")
            turnoverrae = soup.findAll(class_='bitkan-coin-process-title')
            rate = turnoverrae[0].contents[1].string[0:-1]
            timesta = time.strftime('%Y-%m-%d',time.localtime(time.time()))
            sheet = workbook.get_sheet_by_name(coin_list[i])
            max_row = sheet.max_row
            sheet.cell(max_row+1,1,timesta)
            sheet.cell(max_row+1,2,rate)
            workbook.save(path)

#获取贪婪指数
def get_fearAndgreed():
    response = requests.get('https://api.alternative.me/fng/?limit=100000')
    jsonstr = json.loads(response.text)
    fear_greed_list = jsonstr['data']
    # 1、打开文件
    path = 'D:\data\\bitcoin\\fearAndgreed.xls'
    # 2、创建sheet表格
    f = xlwt.Workbook()
    sheet1 = f.add_sheet(r'fearAndgreed', cell_overwrite_ok=True)
    sheet_head = ['日期', '贪婪指数', '贪婪描述']
    #3、写表头
    for i in range(len(sheet_head)):
        sheet1.write(0, i, sheet_head[i])
    for i in range(0,len(fear_greed_list)):
        value = fear_greed_list[i]['value']
        value_classification = fear_greed_list[i]['value_classification']
        timestamp =  time.strftime("%Y-%m-%d", time.localtime(int(fear_greed_list[i]['timestamp'])))
        #4、开始写excel内容
        sheet1.write(i+1,0,timestamp)
        sheet1.write(i+1,1,value)
        sheet1.write(i+1,2,value_classification)
        f.save(path)

get_allBitCoinPrice()
get_allBitCoinGBI()
get_foxOTC()
get_tournover_rate()
get_fearAndgreed()